from flask import Flask, redirect, render_template, request, jsonify
from bs4 import BeautifulSoup
from lxml import etree
import requests
import json
import time
import datetime
import re
import math
from openpyxl import load_workbook, Workbook
from win32com.client import Dispatch
import pythoncom

from yahooquery import Ticker
import numpy as np
import pandas as pd

# using the shitil to copy the excel from local to remote server (10.7.6.199)
import os
import shutil

app = Flask(__name__)

@app.route('/downloadexecl')
def downloadexcel():
    return 'ok'

@app.route('/dutyupdate')
def dutyupdate():
    print('downloading the file')
    dOaction = os.startfile('getexcel.url')
    dst_path = r"\\10.7.6.199\c$\wamp64\www\handover\dutycheck\code\Roster_2020.xlsx"
    print('process the await')
    time.sleep(1)
    print('stop the .199 excel')
    requests.get('http://10.7.6.199:4998/stopExcel')
    time.sleep(1)
    print('start copy to .199')
    shutil.copy(r'C:\Users\09060.gary.wu\Downloads\Roster_2020.xlsx', dst_path)
    time.sleep(3)
    print('run the .199 excel')
    requests.get('http://10.7.6.199:4998/runExcel')
    print('delete the file')
    os.remove(r'C:\Users\09060.gary.wu\Downloads\Roster_2020.xlsx')
    return 'ok'

@app.route('/hm')
def hmfunction():
    return render_template('hmindex.html')

@app.route('/collect')
def collect():
    target_url = 'https://www.gurufocus.com/stock_list.php?m_country[]=USA&m_country[]=_India&m_country[]=IND&m_country[]=PAK&r=USA&p=0&n=30'
    # count_list = list(range(10))
    count_list = list(range(693))
    # for x in count_list:
    #     print(f'https://www.gurufocus.com/stock_list.php?m_country[]=USA&m_country[]=_India&m_country[]=IND&m_country[]=PAK&r=USA&p={x}&n=30')

    user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/51.0.2704.103 Safari/537.36"
    headers = {"User-Agent": user_agent}  #请求头,headers是一个字典类型
    
    stocklist = []
    for cpage in count_list:
        target_url = f'https://www.gurufocus.com/stock_list.php?m_country[]=USA&r=USA&p={cpage}&n=30'
        print(f'CurrentPage is {cpage}')
        try:
            target_html = requests.get(target_url, headers=headers).text
            soup = BeautifulSoup(target_html, 'lxml')
            target_ahref = soup.findAll('a', attrs={'href':re.compile("^/stock/"),'class':'nav'})
            for x in target_ahref:
                if x.text != 'Summary':
                    stocklist.append(dict(text=x.text))
            print(stocklist)
        except Exception as e:
            print(e)
    
    with open('db.json', 'w') as f_write:
        result_json = json.dump(stocklist, f_write)
    f_write.close()
            
    return 'ok'

@app.route('/allstock')
def allstock():
    with open('db.json', 'r') as f_read:
        result = f_read.read()
        result2 = json.loads(result)
    f_read.close()
    return jsonify(result2)

@app.route('/front')
def index():
    return render_template('frontend.html')

@app.route('/stock2/<stock>')
def index3(stock):
    user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/51.0.2704.103 Safari/537.36"
    headers = {"User-Agent": user_agent}  #请求头,headers是一个字典类型
    
    html_statistics = requests.get(f'https://finance.yahoo.com/quote/{stock}/key-statistics?p={stock}', headers=headers).text
    
    soup_for_so = BeautifulSoup(html_statistics, "lxml")
    target = soup_for_so.find('span', text="Shares Outstanding")
    parent_target = target.parent
    target_so = parent_target.next_sibling.text
    print(target_so)
    if 'B' in target_so:
        print('has B')
        so_array = re.findall(r'\d+.\d',target_so)
        so_number_b = so_array[0]
        so_number = float(so_number_b) * 1000
    else:
        print('can not find B')
        so_array = re.findall(r'\d+.\d',target_so)
        so_number = so_array[0]
    print(so_number)
    return 'ok'
    
@app.route('/stock/<stock>')
def index2(stock):
    
    dtime = datetime.datetime.now()
    ans_time = time.mktime(dtime.timetuple())
    
    user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/51.0.2704.103 Safari/537.36"
    headers = {"User-Agent": user_agent}  #请求头,headers是一个字典类型
    
    
    c_list_for_url = ['cash-flow', 'financials', 'analysis']
    
    html_cash = requests.get(f'https://finance.yahoo.com/quote/{stock}/cash-flow?p={stock}', headers=headers).text
    html_financials = requests.get(f'https://finance.yahoo.com/quote/{stock}/financials?p={stock}', headers=headers).text
    html_analysis = requests.get(f'https://finance.yahoo.com/quote/{stock}/analysis?p={stock}', headers=headers).text
    html_statistics = requests.get(f'https://finance.yahoo.com/quote/{stock}/key-statistics?p={stock}', headers=headers).text
    
    cash_date_list = []
    financials_date_list = []
    analysis_date_list = ['2020', '2021']
    cash_data_list = []
    financials_revenue_list = []
    financials_netincome_list = []
    analysis_data_list = []
    temp_error = ""
    return_string = ""
    
    def callbackintM(target):
        x = target.replace(',','')
        a_int = int(x) / 1000
        a_int = '{:g}'.format(a_int)
        return a_int
        
    
    try:
        html_wacc = requests.get(f'https://www.gurufocus.com/term/wacc/{stock}/WACC-', headers=headers).text
        soup_for_wacc = BeautifulSoup(html_wacc, "lxml")
        target_h1 = soup_for_wacc.find("h1")
        target_wacc = target_h1.next_sibling.text
        get_number_only_array = re.findall(r'[0-9]+.',target_wacc)
        wacc_str = "".join(get_number_only_array)
        return_string = return_string + f'Page1 - Personal Required Rate of Return - <a href=https://www.gurufocus.com/term/wacc/{stock}/WACC- target=_blank>wacc</a> (B4): {wacc_str} <br>'
        print(f'Page1 - Personal Required Rate of Return - wacc (B4): {wacc_str}')
    except Exception as e:
        print(e)
        wacc_str = 'None'  
        temp_error = temp_error + f"wacc can't find from <a href=https://www.gurufocus.com/term/wacc/{stock}/WACC- target=_blank>https://www.gurufocus.com/term/wacc/{stock}/WACC-</a><br>"
     
    try:
        soup_for_so = BeautifulSoup(html_statistics, "lxml")
        target = soup_for_so.find('span', text="Shares Outstanding")
        parent_target = target.parent
        target_so = parent_target.next_sibling.text
        if 'B' in target_so:
            so_array = re.findall(r'\d+.\d+',target_so)
            so_number_n = so_array[0]
            so_number = float(so_number_n) * 1000
        else:
            so_array = re.findall(r'\d+.\d+',target_so)
            so_number = so_array[0]
        return_string = return_string + f'Page1 - <a href=https://finance.yahoo.com/quote/{stock}/key-statistics?p={stock} target=_blank>Shares Outstanding</a> (unit:M) - (B5): {so_number} <br>'
        print(f'Page1 - Shares Outstanding (unit:M) - (B5): {so_number}')
    except Exception as e:
        print(e)
        so_number = 0
        temp_error = temp_error + f"Shares Outstanding can't find from <a href=https://finance.yahoo.com/quote/{stock}/key-statistics?p={stock} target=_blank>https://finance.yahoo.com/quote/{stock}/key-statistics?p={stock}</a><br>"
        
     
    for y in c_list_for_url:
        if y == 'cash-flow':
            print('html_cash_start')
            try:
                soup = BeautifulSoup(html_cash, "lxml")
                target_date = soup.find("div", class_= "D(tbhg)").findAll("div")
                target_c_stock_number = soup.find("span", class_= "Trsdu(0.3s) Fw(b) Fz(36px) Mb(-4px) D(ib)").text
                
                # return_string = f'Current stock price: {target_c_stock_number} <br>' + return_string
            
                for x in target_date:
                    cash_date_list.append(x.text)
                del cash_date_list[0:4]
            
            
                target = soup.find("div", title= "Free Cash Flow")
                parent_target = target.parent
                cash_data_list.append(callbackintM(parent_target.next_sibling.next_sibling.text))
                cash_data_list.append(callbackintM(parent_target.next_sibling.next_sibling.next_sibling.text))
                cash_data_list.append(callbackintM(parent_target.next_sibling.next_sibling.next_sibling.next_sibling.text))
                cash_data_list.append(callbackintM(parent_target.next_sibling.next_sibling.next_sibling.next_sibling.next_sibling.text))
            
                return_string = return_string + f'Page1 - <a href=https://finance.yahoo.com/quote/{stock}/cash-flow?p={stock} target=_blank>Free Cash Flow</a> (unit: M) - (B9, C9, D9, E9): <br>'
                print('Page1 - Free Cash Flow (unit: M) - (B9, C9, D9, E9):')
            
                for x, y in zip(cash_date_list, cash_data_list):
                    return_string = return_string + f'{x} {y} <br>'
                    print(x, y)
            except Exception as e:
                print(e)
                temp_error = temp_error + f"Free Cash Flow can't find the data from yahoo page - <a href=https://finance.yahoo.com/quote/{stock}/cash-flow?p={stock} target=_blank>https://finance.yahoo.com/quote/{stock}/cash-flow?p={stock}</a> <br>"
                    
        elif y == 'financials':
            print('html_financials')
            try:
                soup = BeautifulSoup(html_financials, "lxml") # 指定 lxml 作為解析器

                target_date = soup.find("div", class_= "D(tbhg)").findAll("div")

                for y in target_date:
                    financials_date_list.append(y.text)
                del financials_date_list[0:4]    

                target_tr = soup.find("div", title= "Total Revenue")
                parent_target_tr = target_tr.parent
                financials_revenue_list.append(callbackintM(parent_target_tr.next_sibling.next_sibling.text))
                financials_revenue_list.append(callbackintM(parent_target_tr.next_sibling.next_sibling.next_sibling.text))
                financials_revenue_list.append(callbackintM(parent_target_tr.next_sibling.next_sibling.next_sibling.next_sibling.text))
                financials_revenue_list.append(callbackintM(parent_target_tr.next_sibling.next_sibling.next_sibling.next_sibling.next_sibling.text))

                return_string = return_string + f'Page1 - <a href=https://finance.yahoo.com/quote/{stock}/financials?p={stock} target=_blank>Revenue</a> (unit: M) - (B12, C12, D12, E12): <br>'
                print('Page1 - Revenue (unit: M) - (B12, C12, D12, E12): ')
                for x, y in zip(financials_date_list, financials_revenue_list):
                    return_string = return_string + f'{x} {y} <br>'
                    print(x, y)
            except Exception as e:
                print(e)
                temp_error = temp_error + f"revenue can't find the data from yahoo page - <a href=https://finance.yahoo.com/quote/{stock}/financials?p={stock} target=_blank>https://finance.yahoo.com/quote/{stock}/financials?p={stock}</a><br>"
                
            try:    
                target_ni = soup.find("div", title= "Net Income from Continuing & Discontinued Operation")
                parent_target_ni = target_ni.parent
                financials_netincome_list.append(callbackintM(parent_target_ni.next_sibling.next_sibling.text))
                financials_netincome_list.append(callbackintM(parent_target_ni.next_sibling.next_sibling.next_sibling.text))
                financials_netincome_list.append(callbackintM(parent_target_ni.next_sibling.next_sibling.next_sibling.next_sibling.text))
                financials_netincome_list.append(callbackintM(parent_target_ni.next_sibling.next_sibling.next_sibling.next_sibling.next_sibling.text))

                return_string = return_string + f'Page1 - <a href=https://finance.yahoo.com/quote/{stock}/financials?p={stock} target=_blank>Net income</a> (unit: M) - (B13, C13, D13, E13): <br>'
                print('Page1 - Net income (unit: M) - (B13, C13, D13, E13): ')
                for x, y in zip(financials_date_list, financials_netincome_list):
                    return_string = return_string + f'{x} {y} <br>'
                    print(x, y)
            except Exception as e:
                print(e)
                temp_error = temp_error + f"net income can't find the data from yahoo page - <a href=https://finance.yahoo.com/quote/{stock}/financials?p={stock} target=_blank>https://finance.yahoo.com/quote/{stock}/financials?p={stock}</a> <br>"
            
        elif y == 'analysis':
            print('html_analysis')
            soup = BeautifulSoup(html_analysis, "lxml")
            try: 
                # soup = BeautifulSoup(html_analysis, "lxml")

                target = soup.find('span', text="Revenue Estimate")
                parent_target = target.parent.parent.parent
                parent_target_n = parent_target.next_sibling

                for x in parent_target_n.contents:
                    if "Low Estimate" in str(x):
                        for xx in x:
                            print(xx)
                            if 'B' in xx.text:
                                item_array = re.findall(r'\d+.\d+', xx.text)
                                print(item_array)
                                analysis_data_list.append(item_array[0])
                            elif 'M' in xx.text:
                                item_array = re.findall(r'\d+.\d+', xx.text)
                                print(item_array)
                                item_array_b = float(item_array[0]) / 1000
                                analysis_data_list.append(str(item_array_b))
                            else:
                                analysis_data_list.append(xx.text)
                                print(xx.text)
                del analysis_data_list[0:3]

                return_string = return_string + f'Page2 - <a href=https://finance.yahoo.com/quote/{stock}/analysis?p={stock} target=_blank>Analysis</a> - Revenue Estimate - Low Estimate (unit: B) - (D11, E11): <br>'
                print('Page2 - Analysis - Revenue Estimate - Low Estimate (unit: B) - (D11, E11): ')
    
                for x, y  in zip(analysis_date_list, analysis_data_list):
                    return_string = return_string + f'{x} {y} <br>'
                    print(x, y)
            except Exception as e:
                print(e)
                temp_error = temp_error + f"analysis can't find the data from yahoo page - <a href=https://finance.yahoo.com/quote/{stock}/analysis?p={stock} target=_blank>https://finance.yahoo.com/quote/{stock}/analysis?p={stock}</a><br>"
            
            try:
                target_sales = soup.find('span', text="Sales Growth (year/est)")
                parenet_target_sales = target_sales.parent
                sales_thisy = parenet_target_sales.next_sibling.next_sibling.next_sibling.text
                sales_thisy_rstrip = sales_thisy.rstrip('%').strip()
                # sales_thisy_array = re.findall(r'\d+.\d',sales_thisy)
                sales_thisy_number = float(sales_thisy_rstrip)
                # print(f'sales_thisy_number - {sales_thisy}')
                sales_ny = parenet_target_sales.next_sibling.next_sibling.next_sibling.next_sibling.text
                sales_ny_rstrip = sales_ny.rstrip('%').strip()
                # sales_ny_array = re.findall(r'\d+.\d',sales_ny)
                sales_ny_number = float(sales_ny_rstrip)
                # print(f'sales_ny_number - {sales_ny_number}')
                total = round((sales_thisy_number + sales_ny_number)/2, 3)
                total_str = f'{total}%'
            except Exception as e:
                print(e)
                temp_error = temp_error + f"Analysis_Sales Growth (year/est) can't find the data from yahoo page - <a href=https://finance.yahoo.com/quote/{stock}/analysis?p={stock} target=_blank>https://finance.yahoo.com/quote/{stock}/analysis?p={stock}</a><br>"
            
            
                
    wb = load_workbook('sample.xlsx')

    sheet = wb['Step1 - Input Data']

    sheet['B3'].value = stock
    sheet['B4'].value = wacc_str
    sheet['B5'].value = so_number
    # FCFlist = ['B9', 'C9', 'D9', 'E9']
    FCFlist = ['E9', 'D9', 'C9', 'B9']
    # Revlist = ['B12', 'C12', 'D12', 'E12']
    Revlist = ['E12', 'D12', 'C12', 'B12']
    # NetClist = ['B13', 'C13', 'D13', 'E13']
    NetClist = ['E13', 'D13', 'C13', 'B13']
    Analist = ["D11", "E11"]
    try:
        for x, y in zip(FCFlist, cash_data_list):
            # print('cash_data_list')
            # print(x)
            # print(y)
            sheet[x].value = y
            # print(sheet[x].value)
        for x, y in zip(Revlist, financials_revenue_list):
            # print('financials_revenue_list')
            # print(x)
            # print(y)
            sheet[x].value = y
            # print(sheet[x].value)
        for x, y in zip(NetClist, financials_netincome_list):
            # print('financials_netincome_list')
            # print(x)
            # print(y)
            sheet[x].value = y
            # print(sheet[x].value)
    except Exception as e:
        print(e)
    
    zzlist = []
    zz2list = []
     
    try:
        for x, y in zip(financials_netincome_list, financials_revenue_list):
            # print('P2C3')
            # print(x, y)
            zzlist.append(float(x)/float(y))
    except Exception as e:
        print(e)
        
    try:
        for x, y in zip(cash_data_list, financials_netincome_list):
            # print('P2C4')
            # print(x, y)
            zz2list.append(float(x)/float(y))
    except Exception as e:
        print(e)
    
        
    sheet2 = wb['Step2 - Projection']
    
    try:
        for x, y in zip(Analist, analysis_data_list):
            sheet2[x].value = float(y) * 1000
    except Exception as e:
        print(e)
        
        
    try:
        p2c3_float = sum(zzlist) / len(zzlist)
        p2c4_float = sum(zz2list) / len(zz2list)
        # p2c3_float = round(sum(zzlist) / len(zzlist), 2)
        # p2c4_float = round(sum(zz2list) / len(zz2list), 4)
        # p2c3_str = f'{p2c3_float}%'
        # p2c4_str = f'{p2c4_float}%'
        sheet2.cell(row=3, column=3, value=p2c3_float)
        sheet2.cell(row=4, column=3, value=p2c4_float)
        return_string = return_string + f'Page2 - Adopted - Avg Profit Margin (C3): {p2c3_float} <br>'
        return_string = return_string + f'Page2 - Adopted - Avg FCF/ Profit Margin (C4): {p2c4_float} <br>'
        print('Page2 - Adopted - Avg Profit Margin (C3): ', p2c3_float)
        print('Page2 - Adopted - Avg FCF/ Profit Margin (C4): ', p2c4_float)
    except Exception as e:
        print(e)
    

    # print('P2C3_inexcel ', sheet2.cell(row=3, column=3).value)
    # print('P2C4_inexcel ', sheet2.cell(row=4, column=3).value)
    # sheet2['C3'].value = p2c3_str
    # sheet2['C4'].value = p2c4_str
    try:
        return_string = return_string + f'Page2 - Adopted - Growth Rate (C5): {total_str}'
        print(f'Page2 - Adopted - Growth Rate (C5): {total_str}')
        sheet2['C5'].value = total_str
    except Exception as e:
        print(e)
    # print('Growth Rate_inexcel -', sheet2.cell(row=5, column=3).value)

    wb.save(f'static/{stock}.xlsx')
    
    # Fair_Value_of_Equity = round(sheet2_r['B19'].value, 2)
    # print('Fair Value of Equity - ',sheet2['B19'].value)
    wb.close()
    
    
    def just_open(filename):
        print(filename)
        xlApp = Dispatch("Excel.Application")
        xlApp.Visible = False
        xlBook = xlApp.Workbooks.Open(Filename=filename)
        xlBook.Save()
        xlBook.Close()

    try:
        pythoncom.CoInitialize()
        just_open(f'C:/Users/09060.gary.wu/code/FlaskRESTfulAPI/code/static/{stock}.xlsx')

        wb2 = load_workbook(f'C:/Users/09060.gary.wu/code/FlaskRESTfulAPI/code/static/{stock}.xlsx', data_only=True)
        last_sheet = wb2['Step2 - Projection']
        print('Fair_Value_of_Equity')
        print(last_sheet['B19'].value)
        Fair_Value_of_Equity = last_sheet['B19'].value
        return_string = f'Fair_Value_of_Equity: {Fair_Value_of_Equity} <br>' + return_string
        return_string = f'Current stock price: {target_c_stock_number} <br>' + return_string

        wb2.close()
    except Exception as e:
        print(e)
        return_string = f'Current stock price: {target_c_stock_number} <br>' + return_string
    
    # try:
    #     Fair_Value_of_Equity = round(last_sheet['B19'].value, 2)
    #     print('Fair_Value_of_Equity - ',Fair_Value_of_Equity)
    #     print(last_sheet['C3'].value)
    #     print(last_sheet['C4'].value)
    #     wb2.close()
    # except Exception as e:
    #     print(e)
    #     wb2.close()
    #     temp_error = temp_error + f"Fair_Value_of_Equity can't find the data from yahoo page - <a href=https://finance.yahoo.com/quote/{stock}/key-statistics?p={stock} target=_blank>https://finance.yahoo.com/quote/{stock}/key-statistics?p={stock}</a><br>"
            
    list_to_return = []
    list_to_return.append(dict(name=f'Download stock excel - {stock}'))
    list_to_return.append(dict(link=f'/file/{stock}?Cur_stock_value={target_c_stock_number}&time={ans_time}'))
    list_to_return.append(dict(error=f'{temp_error}'))
    list_to_return.append(dict(desc=f'{return_string}'))
    
    if temp_error:
        return f'Get issue when we query the <a href=/file/{stock}?time={ans_time} target=_blank><b>{stock}</b></a><br> {temp_error}'
    else:
        return jsonify(list_to_return)
        # return f'/file/{stock}?Cur_stock_value={target_c_stock_number}&time={ans_time}'
        # return f'/file/{stock}?Cur_stock_value={target_c_stock_number}'
        # return f'/file/{stock}?Cur_stock_value={target_c_stock_number}&Fair_Value_of_Equity={Fair_Value_of_Equity}'
        # return f'/file/{stock}?Cur_stock_value={target_c_stock_number}&Fair_Value_of_Equity={Fair_Value_of_Equity}&time={ans_time}'
                
                
    # return 'ok'
    # return f'/file/{stock}'
    # return app.send_static_file(f'{stock}.xlsx')
    
@app.route('/file/<stock>')
def filestock(stock):
    return app.send_static_file(f'{stock}.xlsx')

@app.route('/vue')
def vue111():
    return render_template('vue.html')

@app.route('/stock_ver2/<target>', methods=['GET', 'POST'])
def testtest(target):
    ## Refer - https://github.com/dpguthrie/yahooquery
    stocks = Ticker(target)
    returndict = lambda a: a[target]
    dict_financial_data = stocks.financial_data
    Current_Price = str(returndict(dict_financial_data)['currentPrice'])
    return_string = ""
    return_string = return_string + f'Current Price: {Current_Price}' + '<br>'
    print(f'Current Price: {Current_Price}')
    
    temp_error = ""

    ## Get WACC
    try:
        user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/51.0.2704.103 Safari/537.36"
        headers = {"User-Agent": user_agent}  #请求头,headers是一个字典类型
        html_wacc = requests.get(f'https://www.gurufocus.com/term/wacc/aapl/WACC-', headers=headers).text
        soup_for_wacc = BeautifulSoup(html_wacc, "lxml")
        target_h1 = soup_for_wacc.find("h1")
        target_wacc = target_h1.next_sibling.text
        get_number_only_array = re.findall(r'[0-9]+.',target_wacc)
        wacc_str = "".join(get_number_only_array)
        return_string = return_string + f'Page1 - Personal Required Rate of Return - wacc (B4): {wacc_str}' + '<br>'
        print(f'Page1 - Personal Required Rate of Return - wacc (B4): {wacc_str}')
    except Exception as e:
        temp_error = temp_error + 'Get issue when query the wacc <br>'
        print('Get issue when query the wacc')
        print(e)

    try:
        dict_key_stats = stocks.key_stats
        Page2_B18 = returndict(dict_key_stats)['sharesOutstanding']
        r_Page2_B18 = str(Page2_B18)
        return_string = return_string + f'Page1 - Shares Outstanding: {r_Page2_B18}' + '<br>'
        print(f'Page1 - Shares Outstanding: {r_Page2_B18}')
    
        f_cash_flow = stocks.cash_flow(frequency='a')
        f_cash_flow_filter = f_cash_flow[['asOfDate','NetIncome','FreeCashFlow']]
    
        df_income_statement = stocks.income_statement(frequency='a')
        df_TotalRevenue = df_income_statement[['asOfDate', 'TotalRevenue']]

        cur_df = pd.merge(f_cash_flow_filter, df_TotalRevenue).drop(index=4)
        cur_df['Net_Profit_Margins'] = cur_df['NetIncome'] / cur_df['TotalRevenue']
        cur_df['FCF_to_Profit_Margins'] = cur_df['FreeCashFlow'] / cur_df['NetIncome']
        Net_Profit_Margins = cur_df['Net_Profit_Margins'].mean()
        FCF_to_Profit_Margins = cur_df['FCF_to_Profit_Margins'].mean()
        r_Net_Profit_Margins = str(Net_Profit_Margins)
        r_FCF_to_Profit_Margins = str(FCF_to_Profit_Margins)
        return_string = return_string + f'Page2 - Avg Profit Margin: {r_Net_Profit_Margins}' + '<br>'
        print(f'Page2 - Avg Profit Margin: {r_Net_Profit_Margins}')
        return_string = return_string + f'Page2 - Avg FCF/ Profit Margin: {r_FCF_to_Profit_Margins}' + '<br>'
        print(f'Page2 - Avg FCF/ Profit Margin: {r_FCF_to_Profit_Margins}')

        ## Starting Page 2
        dict_earnings_trend = stocks.earnings_trend
        list_earnings_trend = returndict(dict_earnings_trend)['trend']

        list_endDate = []
        list_period = []
        list_revenueEstimate_low = []
        list_revenueEstimate_growth_rate = []

        for x in list_earnings_trend:
            temp_endDate = x['endDate'] if x['endDate'] else None
            list_endDate.append(temp_endDate)
            temp_period = x['period'] if x['period'] else None
            list_period.append(temp_period)
            temp_revenueEstimate_low = x['revenueEstimate']['low'] if x['revenueEstimate']['low'] else None
            list_revenueEstimate_low.append(temp_revenueEstimate_low)
            temp_revenueEstimate_growth_rate = x['revenueEstimate']['growth'] if x['revenueEstimate']['growth'] else None
            list_revenueEstimate_growth_rate.append(temp_revenueEstimate_growth_rate)

        df_revenueEstimate_low = pd.DataFrame({
                'endDate': pd.to_datetime(list_endDate),
                'revenueEstimate_low': pd.Series(list_revenueEstimate_low, dtype='float64', index=list_period),
                'revenueEstimate_growth_rate': pd.Series(list_revenueEstimate_growth_rate, dtype='float64', index=list_period)
            }, index=list_period)

        cus_df = pd.DataFrame(df_revenueEstimate_low, index=['0y', '+1y', '+2y', '+3y'])

        ## pd.loc[] -- get dataframe by row (index)
        ## pd.loc[][] -- the second one [] is get the first[] (dataframe) by columns
        Page2_C5_Sales_Growth = (cus_df.loc['0y']['revenueEstimate_growth_rate'] + cus_df.loc['+1y']['revenueEstimate_growth_rate'])/2
        r_Page2_C5_Sales_Growth = str(Page2_C5_Sales_Growth)
        return_string = return_string + f'Page2 - C5 Sales Growth (year/est): {r_Page2_C5_Sales_Growth}' + '\n'
        print(f'Page2 - C5 Sales Growth (year/est): {r_Page2_C5_Sales_Growth}')

        ## target ## 目前要把endDate改成我要的 2020-09-30, 2021-09-30 ... etc
        # 產生pd.Series
        # get the first datetime from dataframe-Analysis
        tempkeydate = cus_df.iloc[0][0]
        # dict for datetime - interval is 1 year
        tempdict = {'endDate': pd.date_range(tempkeydate, periods=4, freq='12M')}
        # new df for temp, it will replace the cur_pd
        df_temp = pd.DataFrame(tempdict, index=['0y', '+1y', '+2y', '+3y'])
        # replace it
        cus_df = cus_df.assign(endDate=df_temp['endDate'])

        ### 加入F11 and G11 (revenue)
        ## F11 = E11*(1+ C5)
        # call E11
        Page2_E11 = cus_df.loc['+1y']['revenueEstimate_low']
        # call F11
        Page2_F11 = Page2_E11 * ( 1+ Page2_C5_Sales_Growth)
        # call G11
        Page2_G11 = Page2_F11 * ( 1+ Page2_C5_Sales_Growth)

        # replace cus_df.iloc[3][3] and [4][4]
        cus_df.loc['+2y', 'revenueEstimate_low'] = Page2_F11
        cus_df.loc['+3y', 'revenueEstimate_low'] = Page2_G11


        ### Page2 _ D12, E12, F12, G12 - Net Income
        ## D12=D11*$C$3
        temp_series = cus_df.loc[:,('revenueEstimate_low',)] * Net_Profit_Margins
        cus_df.insert(3, 'Page2NetIncome', temp_series)

        '''
               endDate  revenueEstimate_low  revenueEstimate_growth_rate  Page2NetIncome
        0y  2020-09-30         2.546180e+11                        0.014    5.469928e+10
        +1y 2021-09-30         2.616590e+11                        0.121    5.621189e+10
        +2y 2022-09-30         2.793210e+11                          NaN    6.000619e+10
        +3y 2023-09-30         2.981752e+11                          NaN    6.405661e+10
        '''

        ## D10 = D12 * C4 (Page2FreeCashFlow = Net Income * FCF_to_Profit_Margins)
        Page2FreeCashFlow_pd_series = cus_df.loc[:,'Page2NetIncome'] * FCF_to_Profit_Margins
        cus_df.insert(4, 'Page2FreeCashFlow', Page2FreeCashFlow_pd_series)

        '''
               endDate  revenueEstimate_low  revenueEstimate_growth_rate  Page2NetIncome  Page2FreeCashFlow
        0y  2020-09-30         2.546180e+11                        0.014    5.469928e+10       5.932015e+10
        +1y 2021-09-30         2.616590e+11                        0.121    5.621189e+10       6.096054e+10
        +2y 2022-09-30         2.793210e+11                          NaN    6.000619e+10       6.507537e+10
        +3y 2023-09-30         2.981752e+11                          NaN    6.405661e+10       6.946796e+10
        '''

        ## D13 = ROUND((1+B15)^1,2) // B15 = wacc_str
        new_wacc_str = wacc_str.replace("%", "")
        wacc_float = float(new_wacc_str)/100
        x_list = list(range(1,5))
        temp_pd_series = pd.Series(x_list, index=['0y', '+1y', '+2y', '+3y'])
        Page2DiscountFactor = pd.DataFrame(round((wacc_float+1)**temp_pd_series,2))
        cus_df.insert(5, 'Page2DiscountFactor', Page2DiscountFactor)

        '''
               endDate  revenueEstimate_low  revenueEstimate_growth_rate  Page2NetIncome  Page2FreeCashFlow  Page2DiscountFactor
        0y  2020-09-30         2.546180e+11                        0.014    5.469928e+10       5.932015e+10                 1.07
        +1y 2021-09-30         2.616590e+11                        0.121    5.621189e+10       6.096054e+10                 1.14
        +2y 2022-09-30         2.793210e+11                          NaN    6.000619e+10       6.507537e+10                 1.22
        +3y 2023-09-30         2.981752e+11                          NaN    6.405661e+10       6.946796e+10                 1.30
        '''

        ## D14 = D10 / D13 ( Page2_PVofFutureCashFlow =  Page2FreeCashFlow / Page2DiscountFactor)
        cus_df['Page2_PVofFutureCashFlow'] = cus_df['Page2FreeCashFlow'] / cus_df['Page2DiscountFactor']

        '''
               endDate  revenueEstimate_low  revenueEstimate_growth_rate  Page2NetIncome  Page2FreeCashFlow  Page2DiscountFactor  Page2_PVofFutureCashFlow
        0y  2020-09-30         2.546180e+11                        0.014    5.469928e+10       5.932015e+10                 1.07              5.543939e+10
        +1y 2021-09-30         2.616590e+11                        0.121    5.621189e+10       6.096054e+10                 1.14              5.347416e+10
        +2y 2022-09-30         2.793210e+11                          NaN    6.000619e+10       6.507537e+10                 1.22              5.334047e+10
        +3y 2023-09-30         2.981752e+11                          NaN    6.405661e+10       6.946796e+10                 1.30              5.343689e+10
        '''

        ### H14 =H10/H13
        ## H10 =G10*(1+B16)/(B15-B16)
        # B15 = wacc_float
        Page2_B16 = 0.02
        Page2_G10 = cus_df.loc['+3y', 'Page2FreeCashFlow']
        Page2_H10 = Page2_G10*(1+Page2_B16)/(wacc_float - Page2_B16)
        Page2_H13 = cus_df.loc['+3y', 'Page2DiscountFactor']
        Page2_H14 = Page2_H10 / Page2_H13

        ### B17 =SUM(D14:H14)
        ## Today's Value = SUM(Page2_PVofFutureCashFlow)
        Page2_B17_front = cus_df['Page2_PVofFutureCashFlow'].sum()
        Page2_B17 = Page2_B17_front + Page2_H14


        ### B18 = Shares Outstanding
        ## Fair Value of Equity = B17/B18
        f = Page2_B17/Page2_B18
        Page2_FairValueofEquity = '%.2f' % f
        r_Page2_FairValueofEquity = str(Page2_FairValueofEquity)
        return_string = return_string + 'Excel_Page_1' + '\n'
        print('Excel_Page_1')
        return_string = return_string + str(cur_df) + '\n'
        print(cur_df)
        return_string = return_string + 'Excel_Page_2' + '\n'
        print('Excel_Page_2')
        return_string = return_string + str(cus_df) + '\n'
        print(cus_df)
        return_string = return_string + f'Fair Value of Equity: {r_Page2_FairValueofEquity}' + '\n'
        print(f'Fair Value of Equity: {r_Page2_FairValueofEquity}')

        html_page1 = (
            cur_df.style
            .set_table_attributes('width="100%"')
            .set_caption("Excel_Page_1")
            .hide_index() # hide the index columns
            .render() # to html format
            )

        html_page2 = (
            cus_df.style
            .set_table_attributes('width="100%"')
            .set_caption("Excel_Page_2")
            .hide_index() # hide the index columns
            .render() # to html format
            )

        total_page = html_page1 + html_page2
    except Exception as e:
        temp_error = temp_error + 'Get issue when using the Yahoo API <br>'
        print('Get issue when using the Yahoo API')
        print(e)
            
    if request.method == 'POST':
        if temp_error:
            return temp_error
        else:
            return_dict = {'Stock_Name': target, 'Current_Price': Current_Price, 'Fair_Value_of_Equity': Page2_FairValueofEquity}
            # return_json = json.dumps(return_dict)
            return jsonify(return_dict)
    else:
        if temp_error:
            return temp_error
        else:
            return f'<h2> Stock - {target} <br>' + f'Current Price: {Current_Price} <br>' + f'Fair Value of Equity: {r_Page2_FairValueofEquity} <br></h2>' + total_page


app.run(host="10.7.6.85", port=4998, debug=True)
